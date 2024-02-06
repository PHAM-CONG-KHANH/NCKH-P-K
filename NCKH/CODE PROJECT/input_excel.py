from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from init import Person

class ExcelWriter:
    def __init__(self, file_path, columns):
        self.file_path = file_path
        self.columns = columns
        if not self._file_exists():
            self._create_new_excel()
        
        
    def _file_exists(self):
        try:
            workbook = load_workbook(self.file_path)
            workbook.close()
            return True
        except FileNotFoundError:
            return False

    def _create_new_excel(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(self.columns)
        workbook.save(self.file_path)
    
    def _set_border(self, sheet):
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = thin_border    

    def add_person (self, person):
        workbook = load_workbook(self.file_path)
        sheet = workbook.active
        sheet.append([getattr(person, attr) for attr in self.columns])
        workbook.save(self.file_path)